from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from django.contrib.auth import get_user_model
from academics.schoolclass.models import SchoolClass
from academics.enrollment.models import Student, Enrollment
from controller.utils import get_academic_year


class Command(BaseCommand):
    help = "Import teachers from an Excel file"

    def add_arguments(self, parser):
        parser.add_argument(
            "file_path",
            type=str,
            help="Path to the Excel file"
        )

    def handle(self, *args, **options):
        file_path = options["file_path"]

        wb = load_workbook(file_path)
        ws = wb.active

        # Read header row
        headers = [cell.value for cell in ws[3]]
        print(headers)

        try:
            name_idx = headers.index("Name")
            class_idx = headers.index("Class")
        except ValueError:
            self.stdout.write(
                self.style.ERROR("Name or Email column not found")
            )
            return

        for row in ws.iter_rows(min_row=4, values_only=True):
            name = row[name_idx]
            class_name = row[class_idx]
            if not class_name or not name:
                continue
            if "ALUMNI" in class_name:
                continue
            self.stdout.write(self.style.SUCCESS(f"Name: {name}, class_name: {class_name}"))

            _class, _ = SchoolClass.objects.get_or_create(name=class_name.strip().lower(), academic_year_id=get_academic_year())
            student = Student.objects.create(
                name=name.strip().title(),
            )
            Enrollment.objects.create(
                student=student,
                school_class=_class,
            )



        self.stdout.write(
            self.style.SUCCESS("Import completed successfully")
        )
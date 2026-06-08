from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from django.contrib.auth import get_user_model
from teacher.teacher.models import Teacher

User = get_user_model()


class Command(BaseCommand):
    help = "Import teachers from an Excel file"

    def add_arguments(self, parser):
        parser.add_argument(
            "file_path",
            type=str,
            help="Path to the Excel file"
        )

    def handle(self, *args, **options):
        file_path = options["file_path"]

        wb = load_workbook(file_path)
        ws = wb.active

        # Read header row
        headers = [cell.value for cell in ws[2]]
        print(headers)

        try:
            name_idx = headers.index("NAME")
            email_idx = headers.index("EMAIL")
        except ValueError:
            self.stdout.write(
                self.style.ERROR("Name or Email column not found")
            )
            return

        for row in ws.iter_rows(min_row=3, values_only=True):
            name = row[name_idx]
            email = row[email_idx]
            if not email:
                continue
            self.stdout.write(self.style.SUCCESS(f"Name: {name}, Email: {email}"))

            if User.objects.filter(email=email).exists():
                self.stdout.write("wxists")

            user = User.objects.create_user(
                username=email,
                email=email,
                first_name=name.split()[0],
                last_name=name.split()[1:],
            )

            Teacher.objects.create(
                user=user,
            )

            self.stdout.write(self.style.SUCCESS(f"Name: {name}, Email: {email} created"))

        self.stdout.write(
            self.style.SUCCESS("Import completed successfully")
        )
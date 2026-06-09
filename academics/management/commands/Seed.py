from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from academics.schoolclass.models import SchoolClass
from academics.subject.models import Subject, SubjectClass
from teacher.teacher.models import Teacher

User = get_user_model()


class Command(BaseCommand):
    help = "Import teachers from an Excel file"

    def add_arguments(self, parser):
        parser.add_argument(
            "file_path",
            type=str,
            help="Path to the Excel file"
        )

    def handle(self, *args, **options):
        file_path = options["file_path"]

        wb = load_workbook(file_path)
        ws = wb.active

        # Read header row
        headers = [cell.value for cell in ws[1]]
        print(headers)

        try:
            name_idx = headers.index("NAME")
            email_idx = headers.index("EMAIL")
            class_idx = headers.index("CLASS")
        except ValueError:
            self.stdout.write(
                self.style.ERROR("Name or Email column not found")
            )
            return

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[name_idx]
            email = row[email_idx]
            _class = row[class_idx]
            if not email or not _class or not name:
                continue
            subject = Subject.objects.get(name=name)
            teacher = Teacher.objects.get(user__username=email)
            sclass = SchoolClass.objects.get(name=_class)

            if not SubjectClass.objects.filter(teacher=teacher, subject=subject, school_class=sclass).exists():
                SubjectClass.objects.create(teacher=teacher, subject=subject, school_class=sclass)
                self.stdout.write(self.style.SUCCESS(f"Name: {name}, Email: {email}, Class: {_class}"))

        self.stdout.write(
            self.style.SUCCESS("Import completed successfully")
        )

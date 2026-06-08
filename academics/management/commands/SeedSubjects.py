from django.contrib.auth.models import Group
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from django.contrib.auth import get_user_model
from academics.subject.models import Subject
User = get_user_model()


class Command(BaseCommand):
    help = "Import teachers from an Excel file"

    def add_arguments(self, parser):
        parser.add_argument(
            "file_path",
            type=str,
            help="Path to the Excel file"
        )

    def handle(self, *args, **options):
        file_path = options["file_path"]

        wb = load_workbook(file_path)
        ws = wb.active

        # Read header row
        headers = [cell.value for cell in ws[3]]
        print(headers)

        try:
            name_idx = headers.index("NAME")
        except ValueError:
            self.stdout.write(
                self.style.ERROR("Name or Email column not found")
            )
            return


        for row in ws.iter_rows(min_row=3, values_only=True):
            name = row[name_idx]
            if not name:
                continue
            self.stdout.write(self.style.SUCCESS(f"Name: {name}"))
            Subject.objects.create(
                name=name,
                code=name,
            )

        self.stdout.write(
            self.style.SUCCESS("Import completed successfully")
        )
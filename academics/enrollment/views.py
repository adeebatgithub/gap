import logging
from datetime import datetime

import openpyxl
from django.contrib import messages
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.db import transaction
from django.db.models.aggregates import Count
from django.db.models.expressions import F
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.views.generic.base import RedirectView
from django.views.generic.detail import SingleObjectMixin
from django.views.generic.edit import FormView
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from academics.enrollment.models import Student, Enrollment
from academics.schoolclass.models import SchoolClass
from attendance.utils import get_all_leafnodes
from controller.consts import BLOOD_GROUP_CHOICES
from controller.mixins import RedirectToDetail
from controller.utils import get_academic_year
from teacher.attendance.models import Attendance, Session
from .forms import EnrollmentForm
from .forms import StudentImportForm

logger = logging.getLogger(__name__)


class EnrollmentListView(PermissionRequiredMixin, ListView):
    permission_required = "academics.view_enrollment"
    model = Enrollment
    template_name = 'academics/enrollments/list.html'
    context_object_name = 'enrollments'

    def get_template_names(self):
        if self.request.htmx:
            return ['academics/enrollments/partial_list.html']
        return super().get_template_names()

    def get_filters(self):
        filters = {"school_class__academic_year__id": get_academic_year(self.request)}
        if search := self.request.GET.get('search'):
            filters['student__name__icontains'] = search

        if class_name := self.request.GET.get('class_name'):
            filters['school_class__name__icontains'] = class_name

        return filters

    def get_queryset(self):
        queryset = Enrollment.objects.filter(**self.get_filters()).select_related(
            'student', 'school_class'
        ).order_by('school_class', 'student__name')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            "classes": SchoolClass.objects.all()
        })
        return context


class EnrollmentDetailView(PermissionRequiredMixin, DetailView):
    permission_required = "academics.view_enrollment"
    model = Enrollment
    template_name = 'academics/enrollments/detail.html'
    context_object_name = 'enrollment'

    def get_session_lookup(self):
        session_totals = (
            Session.objects.filter(
                subject_class__school_class_id=self.object.school_class_id
            )
            .annotate(
                teacher_code=F("subject_class__teacher__code"),
                subject_name=F("subject_class__subject__name"),
            )
            .values("teacher_code", "subject_name")
            .annotate(total=Count("id"))
        )
        session_lookup = {
            (row["teacher_code"], row["subject_name"]): row["total"]
            for row in session_totals
        }
        return session_lookup

    def get_attendance_report(self):
        queryset = (
            Attendance.objects.filter(
                student=self.object.student,
                status=Attendance.PRESENT,
                session__subject_class__school_class__academic_year=get_academic_year(self.request),
            )
            .annotate(
                teacher_code=F("session__subject_class__teacher__code"),
                subject_name=F("session__subject_class__subject__name"),
            )
            .values("teacher_code", "subject_name")
            .annotate(present_count=Count("id"))
        )

        data = {
            "subjects": {"zTotal", "zz%"},
            "counts": {"out of": {"zTotal": 0, "zz%": 100}, "subjects": {"zTotal": 0, "zz%": 0}}
        }
        session_lookup = self.get_session_lookup()
        for row in queryset:
            subject_name = row["subject_name"]
            teacher_code = row["teacher_code"]
            subject_full = f"{subject_name} ({teacher_code})"
            present_count = row["present_count"]

            data["subjects"].add(subject_full)

            data["counts"]["subjects"][subject_full] = present_count
            data["counts"]["subjects"]["zTotal"] += present_count

            session_total = session_lookup.get(
                (teacher_code, subject_name),
                0,
            )
            data["counts"]["out of"][subject_full] = session_total
            data["counts"]["out of"]["zTotal"] += session_total

        if data["counts"]["out of"]["zTotal"] > 0:
            data["counts"]["subjects"]["zz%"] = round(
                (data["counts"]["subjects"]["zTotal"] / data["counts"]["out of"]["zTotal"]) * 100
            )
        data["subjects"] = sorted(data["subjects"])

        return data

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context.update({
            "attendance_report": self.get_attendance_report(),
        })
        return context


class EnrollmentCreateView(PermissionRequiredMixin, CreateView):
    permission_required = "academics.add_enrollment"
    model = Student
    form_class = EnrollmentForm
    template_name = 'academics/enrollments/form.html'
    success_url = reverse_lazy('academics:enrollment:list')

    def get_initial(self):
        initial = super().get_initial()
        if self.request.GET.get('schoolclass'):
            initial.update({
                "school_class": SchoolClass.objects.get(id=self.request.GET.get('schoolclass')),
                "admission_date": timezone.localdate(),
            })
        return initial

    def get_success_url(self):
        if self.request.GET.get('schoolclass'):
            return reverse_lazy('academics:schoolclass:detail', kwargs={'pk': self.request.GET.get('schoolclass')})
        return super().get_success_url()

    def form_valid(self, form):
        with transaction.atomic():
            self.object = form.save()
            Enrollment.objects.create(
                school_class=form.cleaned_data['school_class'],
                student=self.object,
            )
        return redirect(self.get_success_url())


class EnrollmentUpdateView(PermissionRequiredMixin, RedirectToDetail, UpdateView):
    permission_required = "academics.change_enrollment"
    model = Student
    form_class = EnrollmentForm
    template_name = 'academics/enrollments/form.html'
    success_url = reverse_lazy('academics:enrollment:list')

    def get_initial(self):
        initial = super().get_initial()
        initial.update({
            "school_class": SchoolClass.objects.get(id=Enrollment.objects.get(student=self.object).school_class.id),
        })
        return initial

    def get_detail_url(self):
        return reverse_lazy('academics:enrollment:detail', kwargs={'pk': self.object.pk})

    def get_success_url(self):
        if self.request.GET.get('schoolclass'):
            return reverse_lazy('academics:schoolclass:detail', kwargs={'pk': self.request.GET.get('schoolclass')})
        return super().get_success_url()

    def form_valid(self, form):
        with transaction.atomic():
            self.object = form.save()
            Enrollment.objects.filter(
                student=self.object,
            ).update(school_class=form.cleaned_data['school_class'])
        return redirect(self.get_success_url())


class EnrollmentDeleteView(PermissionRequiredMixin, DeleteView):
    permission_required = "academics.delete_enrollment"
    http_method_names = ('post',)
    model = Enrollment
    success_url = reverse_lazy('academics:enrollment:list')

    def get_detail_url(self):
        return reverse_lazy('academics:enrollment:detail', kwargs={'pk': self.object.pk})

    def get_success_url(self):
        if self.request.GET.get('schoolclass'):
            return reverse_lazy('academics:schoolclass:detail', kwargs={'pk': self.request.GET.get('schoolclass')})
        return super().get_success_url()


class EnrollmentChangeLeaveStatusView(PermissionRequiredMixin, SingleObjectMixin, RedirectView):
    permission_required = "academics.change_enrollment"
    model = Enrollment

    def get_redirect_url(self, *args, **kwargs):
        if self.request.GET.get('usr') and self.request.GET.get('usr') == 'teacher':
            return reverse_lazy('teacher:schoolclass:detail', kwargs={'pk': self.object.school_class.id})
        return reverse_lazy('academics:schoolclass:detail', kwargs={'pk': self.object.school_class.id})

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.on_leave = not self.object.on_leave
        self.object.save()
        return super().get(request, args, kwargs)


class StudentImportView(FormView):
    form_class = StudentImportForm

    EXPECTED_HEADERS = [
        "name",
        "gender",
        "reg number",
        "father name",
        "father phone",
        "mother name",
        "address",
        "email",
        "phone",
        "dob",
        "health issues",
        "blood group",
        "admission date",
    ]

    def get_success_url(self):
        return reverse_lazy("academics:schoolclass:detail", kwargs={"pk": self.kwargs["pk"]})

    def form_invalid(self, form):
        return redirect(self.get_success_url())

    def get_school_class(self):
        try:
            school_class = SchoolClass.objects.get(pk=self.kwargs["pk"])
        except SchoolClass.DoesNotExist:
            raise ValueError("Invalid class selected.")

        if school_class not in get_all_leafnodes():
            raise ValueError("Importing is not allowed")

        return school_class

    @transaction.atomic
    def form_valid(self, form):
        excel_file = form.cleaned_data["file"]

        try:
            self.school_class = self.get_school_class()
        except ValueError as e:
            messages.error(self.request, str(e))
            return self.form_invalid(form)

        try:
            workbook = openpyxl.load_workbook(
                excel_file,
                data_only=True,
                read_only=True,
            )
        except Exception:
            messages.error(
                self.request,
                "Invalid Excel file. Please upload a valid .xlsx file.",
            )
            logger.exception("Failed to read uploaded workbook")
            return self.form_invalid(form)

        sheet = workbook.active

        try:
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows)
        except StopIteration:
            messages.error(self.request, "The uploaded file is empty.")
            return self.form_invalid(form)

        header_row = [
            str(cell).strip().lower() if cell is not None else ""
            for cell in header_row
        ]

        missing_headers = [
            header
            for header in self.EXPECTED_HEADERS
            if header not in header_row
        ]

        if missing_headers:
            messages.error(
                self.request,
                f"Missing required columns: {', '.join(missing_headers)}",
            )
            return self.form_invalid(form)

        created_count = 0
        updated_count = 0
        error_rows = []

        for row_number, row in enumerate(rows, start=2):
            row_data = dict(zip(header_row, row))

            try:
                student_data = self.clean_row(row_data)

                if not student_data["name"]:
                    raise ValueError("'name' is required")

                student, created = self.create_or_update_student(
                    student_data
                )

                Enrollment.objects.get_or_create(
                    student=student,
                    school_class=self.school_class,
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

            except Exception as e:
                logger.exception(
                    "Student import failed for row %s",
                    row_number,
                )
                error_rows.append(
                    f"Row {row_number}: {str(e)}"
                )

        summary = (
            f"Import complete: "
            f"{created_count} created, "
            f"{updated_count} updated."
        )

        if error_rows:
            summary += f" {len(error_rows)} row(s) had errors."

            messages.warning(self.request, summary)

            for error in error_rows[:10]:
                messages.error(self.request, error)

            if len(error_rows) > 10:
                messages.warning(
                    self.request,
                    f"...and {len(error_rows) - 10} more errors."
                )
        else:
            messages.success(self.request, summary)

        return super().form_valid(form)

    def create_or_update_student(self, student_data):
        """
        Prefer reg_number as unique identifier.
        Fall back to name only if reg_number is not available.
        """

        reg_number = student_data.get("reg_number")

        if reg_number:
            return Student.objects.update_or_create(
                reg_number=reg_number,
                defaults=student_data,
            )

        return Student.objects.update_or_create(
            name=student_data["name"],
            defaults=student_data,
        )

    def clean_row(self, row_data):
        gender_map = {
            "male": "M",
            "m": "M",
            "female": "F",
            "f": "F",
        }

        valid_blood_groups = {
            choice[0]
            for choice in (BLOOD_GROUP_CHOICES or [])
        }

        def parse_date(value):
            if value in (None, ""):
                return None

            if isinstance(value, datetime):
                return value.date()

            if hasattr(value, "year"):
                return value

            value = str(value).strip()

            formats = (
                "%Y-%m-%d",
                "%d/%m/%Y",
                "%d-%m-%Y",
                "%m/%d/%Y",
            )

            for fmt in formats:
                try:
                    return datetime.strptime(
                        value,
                        fmt,
                    ).date()
                except ValueError:
                    continue

            raise ValueError(
                f"Invalid date format: {value}"
            )

        gender_raw = (
            str(row_data.get("gender") or "")
            .strip()
            .lower()
        )

        if gender_raw:
            if gender_raw not in gender_map:
                raise ValueError(
                    f"Invalid gender: {gender_raw}"
                )
            gender = gender_map[gender_raw]
        else:
            gender = "M"

        blood_group_raw = (
            str(row_data.get("blood group") or "")
            .strip()
            .upper()
        )

        if blood_group_raw:
            if blood_group_raw not in valid_blood_groups:
                raise ValueError(
                    f"Invalid blood group: {blood_group_raw}"
                )
            blood_group = blood_group_raw
        else:
            blood_group = None

        return {
            "name": str(
                row_data.get("name") or ""
            ).strip(),
            "gender": gender,
            "reg_number": (
                    str(row_data.get("reg number") or "").strip()
                    or None
            ),
            "father_name": (
                    str(row_data.get("father name") or "").strip()
                    or None
            ),
            "father_phone": (
                    str(row_data.get("father phone") or "").strip()
                    or None
            ),
            "mother_name": (
                    str(row_data.get("mother name") or "").strip()
                    or None
            ),
            "address": (
                    str(row_data.get("address") or "").strip()
                    or None
            ),
            "email": (
                    str(row_data.get("email") or "").strip()
                    or None
            ),
            "phone": (
                    str(row_data.get("phone") or "").strip()
                    or None
            ),
            "dob": parse_date(
                row_data.get("dob")
            ),
            "health_issues": (
                    str(row_data.get("health issues") or "").strip()
                    or None
            ),
            "blood_group": blood_group,
            "admission_date": parse_date(
                row_data.get("admission date")
            ),
        }


class StudentImportTemplateView(View):
    HEADERS = [
        ("name", "Name"),
        ("gender", "Gender"),
        ("reg_number", "Reg Number"),
        ("father_name", "Father Name"),
        ("father_phone", "Father Phone"),
        ("mother_name", "Mother Name"),
        ("address", "Address"),
        ("email", "Email"),
        ("phone", "Phone"),
        ("dob", "DOB"),
        ("health_issues", "Health Issues"),
        ("blood_group", "Blood Group"),
        ("admission_date", "Admission Date"),
    ]

    SAMPLE_ROW = [
        "John Doe", "M", "2024-CS-001", "Richard Doe", "9876543210",
        "Jane Doe", "123 Main St, Springfield", "john@example.com", "9876500000",
        "2008-05-14", "None", "O+", "2024-06-01",
    ]

    def get(self, request, *args, **kwargs):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Students"

        header_font = Font(bold=True, color="FFFFFF", name="Arial")
        header_fill = PatternFill("solid", start_color="1F2937")
        thin_border = Border(*([Side(style="thin", color="D1D5DB")] * 4))
        center = Alignment(horizontal="center", vertical="center")

        for col_idx, (_, label) in enumerate(self.HEADERS, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        for col_idx, value in enumerate(self.SAMPLE_ROW, start=1):
            cell = sheet.cell(row=2, column=col_idx, value=value)
            cell.font = Font(italic=True, color="6B7280", name="Arial")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

        widths = [20, 10, 15, 20, 15, 20, 30, 25, 15, 18, 20, 12, 20]
        for col_idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        sheet.freeze_panes = "A2"

        gender_dv = DataValidation(type="list", formula1='"M,F"', allow_blank=True)
        sheet.add_data_validation(gender_dv)
        gender_dv.add(f"B3:B1000")

        blood_groups = '"A+,A-,B+,B-,AB+,AB-,O+,O-"'
        blood_dv = DataValidation(type="list", formula1=blood_groups, allow_blank=True)
        sheet.add_data_validation(blood_dv)
        blood_dv.add(f"L3:L1000")

        notes = wb.create_sheet("Instructions")
        notes_font = Font(name="Arial", size=11)
        instructions = [
            "Instructions",
            "",
            "1. Do not change the column headers in the 'Students' sheet.",
            "2. Row 2 contains a sample entry — replace or delete it before importing.",
            "3. 'Name' is the only required field.",
            "4. Gender accepts M or F only.",
            "5. Blood Group accepts A+, A-, B+, B-, AB+, AB-, O+, O- only.",
            "6. Dates must be in YYYY-MM-DD format (e.g. 2024-06-01).",
            "7. If 'Reg Number' matches an existing student, that student's record will be updated instead of duplicated.",
        ]
        for row_idx, line in enumerate(instructions, start=1):
            cell = notes.cell(row=row_idx, column=1, value=line)
            cell.font = Font(bold=True, size=14, name="Arial") if row_idx == 1 else notes_font
        notes.column_dimensions["A"].width = 90

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="student_import_template.xlsx"'
        wb.save(response)
        return response
